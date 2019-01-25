import pickle

import openpyxl


def load_workbook(workbook_name):
    """
    Load an Excel document from disk.

    :param workbook_name:
    :return:
    """
    try:
        with open(workbook_name + '.p', 'rb') as f:
            workbook =  pickle.load(f)
    except Exception as e:
        workbook = openpyxl.load_workbook(workbook_name)
        with open(workbook_name + '.p', 'wb') as f:
            pickle.dump(workbook, f)
    return workbook


class Dimension:
    def __init__(self, min_col, min_row, max_col, max_row):
        self._min_col = min_col
        self._min_row = int(min_row)
        self._max_col = max_col
        self._max_row = int(max_row)

    @property
    def min_col(self):
        return self._min_col

    @min_col.setter
    def min_col(self, x):
        self._min_col = x

    @property
    def min_row(self):
        return self._min_row

    @min_row.setter
    def min_row(self, x):
        self._min_row = x

    @property
    def max_col(self):
        return self._max_col

    @property
    def max_col_num(self):
        return int(ord(str(self._max_col).lower()) - 96)

    @max_col.setter
    def max_col(self, x):
        self._max_col = x

    @property
    def max_row(self):
        return self._max_row

    @max_row.setter
    def max_row(self, x):
        self._max_row = x

    def __repr__(self):
        return '{}{}:{}{}'.format(self.min_col, self.min_row, self.max_col, self.max_row)


def get_dimensions(worksheet):
    """
    Get the dimensions of the worksheet.

    :param worksheet: The excel worksheet
    :return: A Dimension object containing properties: min_col, min_row, max_col, max_row
    :rtype: Dimension
    """
    dimension = worksheet.calculate_dimension()
    min_col = []
    min_row = []
    for i, letter in enumerate(dimension):
        if letter == ':':
            end = i+1
            break
        if letter.isalpha():
            min_col.append(letter)
        else:
            min_row.append(letter)
    min_col = ''.join(min_col)
    min_row = ''.join(min_row)

    max_col = []
    max_row = []
    for i, letter in enumerate(dimension[end:]):
        if letter.isalpha():
            max_col.append(letter)
        else:
            max_row.append(letter)
    max_col = ''.join(max_col)
    max_row = ''.join(max_row)

    return Dimension(min_col, min_row, max_col, max_row)
